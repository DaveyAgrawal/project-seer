#!/usr/bin/env python3
"""Convert co2_source_scorecard.csv -> co2_source_scorecard.xlsx (formatted)."""
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

CSV = "/Users/devanagrawal/Desktop/project-seer/map/ingest/co2_source_scorecard.csv"
XLSX = "/Users/devanagrawal/Desktop/project-seer/map/ingest/co2_source_scorecard.xlsx"

# Columns that should be stored as numbers in Excel
INT_COLS = {"emissions_tpy_2023", "site_depth_to_150c_m", "site_basement_depth_m", "transmission_kv"}
FLOAT_COLS = {"mw_potential", "source_lat", "source_lon", "dist_to_viable_km",
              "site_lat", "site_lon", "site_temp_c_at_5km", "site_water_sui",
              "dist_to_co2_pipeline_km", "nearest_transmission_km", "composite_score"}
BOOL_COLS = {"emissions_estimated", "viable_within_50km", "within_30km_co2_pipeline"}


def cast(col, val):
    if val == "":
        return None
    if col in BOOL_COLS:
        return val == "True"
    if col in INT_COLS:
        try:
            return int(float(val))
        except ValueError:
            return val
    if col in FLOAT_COLS:
        try:
            return float(val)
        except ValueError:
            return val
    return val


rows = list(csv.reader(open(CSV)))
header, data = rows[0], rows[1:]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "CO2 Source Scorecard"

# header
hfill = PatternFill("solid", fgColor="1F4E78")
hfont = Font(bold=True, color="FFFFFF")
for c, name in enumerate(header, 1):
    cell = ws.cell(row=1, column=c, value=name)
    cell.fill = hfill
    cell.font = hfont
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# data
for r, row in enumerate(data, 2):
    for c, (col, val) in enumerate(zip(header, row), 1):
        ws.cell(row=r, column=c, value=cast(col, val))

# widths (cap so it stays readable)
for c, name in enumerate(header, 1):
    maxlen = max([len(name)] + [len(row[c - 1]) for row in data]) + 2
    ws.column_dimensions[get_column_letter(c)].width = min(maxlen, 42)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(XLSX)
print(f"Wrote {XLSX}  ({len(data)} rows x {len(header)} cols)")
